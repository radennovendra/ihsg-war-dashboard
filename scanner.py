import os
import time
import pandas as pd
from datetime import datetime

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from signals import evaluate
from backtest import hedge_expectancy
from flow_engine.foreign_stock import stock_foreign_map, classify
from flow_engine.fundamental_engine import get_fundamental
from utils.yahoo_pro import download_price
from utils.rate_guard import guard
from utils.safe_loop import memory_guard
from telegram_engine import send, send_photo, send_file
from utils.chart_generator import generate_chart
from utils.beifraksi import tick_size, floor_tick, ceil_tick

WATCHLIST_TOPN = 15
BATCH_LIMIT = 200

price_cache = {}

def get_market_context(ihsg_df):

    if ihsg_df is None or ihsg_df.empty:
        return "UNKNOWN", "UNKNOWN"

    df = ihsg_df.copy()

    df["MA20"] = df["Close"].rolling(20).mean()
    df["MA50"] = df["Close"].rolling(50).mean()

    last = df.iloc[-1]

    trend = "SIDEWAYS"

    if last["Close"] > last["MA20"] > last["MA50"]:
        trend = "UPTREND"

    elif last["Close"] < last["MA20"] < last["MA50"]:
        trend = "DOWNTREND"

    # Market regime
    regime = "NEUTRAL"

    if trend == "UPTREND":
        regime = "RISK-ON"

    elif trend == "DOWNTREND":
        regime = "RISK-OFF"

    return regime, trend

def build_telegram_message(results, market_regime, ihsg_trend):

    msg = f"""
📊 IHSG SMART MONEY
🕒 {datetime.now().strftime("%d %b %H:%M")}

🌏 Market Regime
IHSG Trend : {ihsg_trend}
Risk Mode  : {market_regime}

━━━━━━━━━━━━━━━━
"""

    medals = ["🥇","🥈","🥉","4️⃣","5️⃣"]

    for i,(sym,r) in enumerate(results[:5]):

        sector = r.get("sector","Unknown")

        entry = f"{int(r['entry_low'])} — {int(r['entry_high'])}"
        tp = int(r["tp2"])
        sl = int(r["stoploss"])

        win20 = r.get("win20d",0)*100
        exp20 = r.get("exp20",0)*100

        foreign = r.get("foreign_net",0)

        if abs(foreign) >= 1e9:
            ftxt = f"{foreign/1e9:.1f}B"
        elif abs(foreign) >= 1e6:
            ftxt = f"{foreign/1e6:.1f}M"
        elif abs(foreign) >= 1e3:
            ftxt = f"{foreign/1e3:.0f}K"
        else:
            ftxt = str(int(foreign))

        medal = medals[i]

        msg += f"""
{medal} {sym} | {sector}
Score : {r['score']}

🎯 Entry : {entry}
🛑 SL    : {sl}
🏁 TP    : {tp}

📈 WinRate20D : {win20:.0f}%
⚡ Exp20D     : {exp20:.2f}%

🌍 Foreign Flow
{ftxt} | {r.get("foreign_status","")}

━━━━━━━━━━━━━━━━
"""

    return msg

def safe_download(sym):

    if sym in price_cache:
        return price_cache[sym]

    for _ in range(3):
        try:
            df = download_price(sym)

            if df is not None and not df.empty:

                if len(price_cache) > 250:
                    price_cache.clear()

                price_cache[sym] = df
                return df
        except:
            pass

        time.sleep(1)

    return None

# ==========================
# SECTOR MAP
# ==========================
def load_sector_map():
    path = "data/sector_map.csv"
    if not os.path.exists(path):
        return {}
    df = pd.read_csv(path)
    return dict(zip(df["Ticker"], df["Sector"]))

SECTOR_MAP = load_sector_map()

# ==========================
# STAR
# ==========================
def star(score):
    if score >= 95: return "⭐⭐⭐⭐⭐"
    if score >= 85: return "⭐⭐⭐⭐"
    if score >= 70: return "⭐⭐⭐"
    if score >= 50: return "⭐⭐"
    return "⭐"

# ==========================
# TOP FOREIGN
# ==========================
def get_top_foreign(results, n=30):
    return sorted(
        results,
        key=lambda x: abs(x[1].get("foreign_net",0)),
        reverse=True
    )[:n]

# ==========================
# EXCEL EXPORT
# ==========================
def export_terminal_excel(results, total_foreign_today, top_foreign, market_regime, ihsg_trend):

    os.makedirs("reports", exist_ok=True)
    file = "reports/HEDGEFUND_TERMINAL.xlsx"
    wb = Workbook()

    def autosize(ws):
        for col in ws.columns:
            length = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = length + 2

    # ==========================
    # MARKET REGIME AUTO
    # ==========================
    regime = market_regime

    if regime == "RISK-ON":
        insight = "IHSG Uptrend. Institutional Risk Appetite Strong."
    elif regime == "RISK-OFF":
        insight = "IHSG Downtrend. Defensive Positioning."
    else:
        insight = "IHSG Sideways. Selective Accumulation"

    # ==========================
    # DASHBOARD – FUND MANAGER VERSION
    # ==========================
    ws = wb.active
    ws.title = "DASHBOARD"

    

    # ===== STYLE =====
    title_font = Font(size=20, bold=True)
    big_font = Font(size=14, bold=True)
    label_font = Font(size=11, bold=True)

    blue = PatternFill("solid", fgColor="1F4E78")
    light = PatternFill("solid", fgColor="EEF3F8")
    green = PatternFill("solid", fgColor="C6EFCE")
    red = PatternFill("solid", fgColor="FFC7CE")
    yellow = PatternFill("solid", fgColor="FFF2CC")

    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    center = Alignment(horizontal="center")

    # ===== TITLE =====
    ws.merge_cells("A1:F1")
    ws["A1"] = "IHSG HEDGE FUND TERMINAL"
    ws["A1"].font = title_font

    ws.merge_cells("A2:F2")
    ws["A2"] = datetime.now().strftime("%d %B %Y")
    ws["A2"].alignment = center

    # ===== CORE METRICS =====
    avg_score = sum(r["score"] for _,r in results)/len(results)
    leaders = sum(1 for _,r in results if r["score"]>=95)

    ws["A4"] = "Average Score"
    ws["A4"].font = label_font
    ws["B4"] = round(avg_score,1)
    ws["B4"].font = big_font

    ws["A5"] = "Leaders"
    ws["A5"].font = label_font
    ws["B5"] = leaders
    ws["B5"].font = big_font

    ws["A6"] = "Total Stocks"
    ws["A6"].font = label_font
    ws["B6"] = len(results)

    # ===== MARKET REGIME =====
    ws["D4"] = "Market Regime"
    ws["D4"].font = label_font
    ws["E4"] = regime
    ws["E4"].font = big_font

    if regime == "RISK-ON":
        ws["E4"].fill = green
    elif regime == "RISK-OFF":
        ws["E4"].fill = red
    else:
        ws["E4"].fill = yellow

    ws.merge_cells("D5:F6")
    ws["D5"] = insight

    # ===== TOTAL FOREIGN =====
    total_foreign_market = load_foreign_today()
    total_foreign_watchlist = sum(r.get("foreign_net",0) for _,r in results)

    def money(x):
        if abs(x)>=1e12: return f"{x/1e12:.2f}T"
        if abs(x)>=1e9: return f"{x/1e9:.1f}B"
        if abs(x)>=1e6: return f"{x/1e6:.1f}M"
        return str(int(x))

    ws["A8"] = "Total Foreign Today (Market)"
    ws["A8"].font = label_font
    ws["B8"] = money(total_foreign_market)
    ws["B8"].font = big_font

    if total_foreign_market>0:
        ws["B8"].fill = green
    else:
        ws["B8"].fill = red
    
    ws["A9"] = "Total Foreign Accum (Top Institutional Picks)"
    ws["A9"].font = label_font
    ws["B9"] = money(total_foreign_watchlist)
    ws["B9"].font = big_font
    
    if total_foreign_watchlist>0:
        ws["B9"].fill = green
    else:
        ws["B9"].fill = red
    
    # ===== TOP SECTOR FLOW =====
    sector_flow = {}
    for sym,r in results:
        sec = r.get("sector","Unknown")
        sector_flow.setdefault(sec,0)
        sector_flow[sec]+=r.get("foreign_net",0)

    top_sector_flow = sorted(sector_flow.items(), key=lambda x:x[1], reverse=True)[:5]

    ws["A10"] = "Top Foreign Sectors"
    ws["A10"].font = label_font

    row=11
    for i,(sec,val) in enumerate(top_sector_flow,1):
        ws[f"A{row}"] = f"{i}. {sec}"
        ws[f"B{row}"] = money(val)
        row+=1
    # ===== TOP SECTOR WINNERS =====
    ws["A16"] = "Top Sector Winners"
    ws["A16"].font = label_font

    sector_strength = {}

    for sym, r in results:
        sec = r.get("sector", "Unknown")

        if sec not in sector_strength:
            sector_strength[sec] = {
                "scores": [],
                "foreign": 0
            }

        sector_strength[sec]["scores"].append(r["score"])
        sector_strength[sec]["foreign"] += r.get("foreign_net", 0)

    # hitung avg score + total foreign
    sector_rank = []

    for sec, data in sector_strength.items():
        avg_score = sum(data["scores"]) / len(data["scores"])
        foreign = data["foreign"]

        sector_rank.append((sec, avg_score, foreign))

    sector_rank.sort(key=lambda x: x[1], reverse=True)

    row = 17

    for i, (sec, avg, foreign) in enumerate(sector_rank[:5], 1):

        if abs(foreign) >= 1e9:
            ftxt = f"{foreign/1e9:.1f}B"
        elif abs(foreign) >= 1e6:
            ftxt = f"{foreign/1e6:.1f}M"
        else:
            ftxt = str(int(foreign))

        ws[f"A{row}"] = f"{i}. {sec}"
        ws[f"B{row}"] = round(avg,1)
        ws[f"C{row}"] = ftxt

        row += 1

    # ===== TOP LEADERS PREVIEW =====
    ws["D8"] = "Top Institutional Leaders"
    ws["D8"].font = label_font

    leaders_sorted = sorted(results, key=lambda x:x[1]["score"], reverse=True)[:5]

    row=9
    for sym,r in leaders_sorted:
        ws[f"D{row}"] = sym
        ws[f"E{row}"] = r["score"]
        ws[f"F{row}"] = money(r.get("foreign_net",0))
        row+=1

    # ===== RISK GAUGE =====
    risk_score = sum(1 for _,r in results if r["score"]<50)/len(results)

    ws["D14"] = "Market Risk"
    ws["D14"].font = label_font

    if risk_score>0.5:
        txt="HIGH RISK"
        fill=red
    elif risk_score>0.3:
        txt="MEDIUM"
        fill=yellow
    else:
        txt="LOW RISK"
        fill=green

    ws["E14"] = txt
    ws["E14"].fill = fill
    ws["E14"].font = big_font

    # ===== WIDTH =====
    ws.column_dimensions["A"].width=26
    ws.column_dimensions["B"].width=18
    ws.column_dimensions["C"].width=5
    ws.column_dimensions["D"].width=26
    ws.column_dimensions["E"].width=18
    ws.column_dimensions["F"].width=18

    autosize(ws)

    # ==========================
    # WATCHLIST
    # ==========================
    ws = wb.create_sheet("WATCHLIST")
    ws.append([
        "Rank","Ticker","Score","Entry","SL","TP1","TP2","TP3",
        "Win20D","Exp20D","PF20","Foreign","Status"
    ])

    for i,(sym,r) in enumerate(results[:20],1):

        win20 = r.get("win20d",0)*100
        exp20 = r.get("exp20",0)*100
        net = r.get("foreign_net",0)

        ws.append([
            i,
            sym,
            r["score"],
            f"{r['entry_low']:.0f}-{r['entry_high']:.0f}",
            r["stoploss"],
            r["tp1"],
            r["tp2"],
            r["tp3"],
            f"{win20:.0f}%",
            f"{exp20:.2f}%",
            round(r.get("pf20",1),2),
            net,
            r.get("foreign_status","")
        ])

    autosize(ws)

    # ==========================
    # FOREIGN FLOW
    # ==========================
    ws = wb.create_sheet("FOREIGN_FLOW")
    ws.append(["Rank","Ticker","ForeignNet"])

    foreign_rank = sorted(results, key=lambda x:abs(x[1].get("foreign_net",0)), reverse=True)

    for i,(sym,r) in enumerate(foreign_rank[:30],1):
        ws.append([i,sym,r.get("foreign_net",0)])

    autosize(ws)

    # ==========================
    # UNIVERSE
    # ==========================
    ws = wb.create_sheet("UNIVERSE")
    ws.append(["Ticker","Score","RawScore","Discount"])

    for sym,r in results:
        disc = r.get("discount_52w",0)
        ws.append([sym,r["score"],r["raw_score"],f"{disc*100:.1f}%"])

    autosize(ws)
    
    # ==========================
    # TOP WINNER PER SECTOR
    # ==========================
    ws = wb.create_sheet("TOP_SECTOR_WINNERS")

    ws.append([
        "Sector",
        "Rank",
        "Ticker",
        "Score",
        "Entry",
        "TP",
        "Foreign"
    ])

    sector_groups = {}

    # group by sector
    for sym, r in results:
        sec = r.get("sector", "Unknown")
        sector_groups.setdefault(sec, []).append((sym, r))

    row = 2

    for sec, stocks in sector_groups.items():

        # sort by score
        stocks_sorted = sorted(stocks, key=lambda x: x[1]["score"], reverse=True)[:5]

        first = True

        for rank, (sym, r) in enumerate(stocks_sorted, 1):

            entry = f"{r['entry_low']:.0f}-{r['entry_high']:.0f}"
            tp = f"{r['tp2']:.0f}"
            net = r.get("foreign_net", 0)

            if abs(net) >= 1e9:
                txt = f"{net/1e9:.1f}B"
            elif abs(net) >= 1e6:
                txt = f"{net/1e6:.1f}M"
            else:
                txt = str(int(net))

            if first:
                ws.append([sec, rank, sym, r["score"], entry, tp, txt])
                first = False
            else:
                ws.append(["", rank, sym, r["score"], entry, tp, txt])

    autosize(ws)
# ==========================
# TOP FOREIGN EXCEL
# ==========================
    ws2 = wb.create_sheet("TOP FOREIGN")

    ws2.append(["Rank","Stock","Foreign Net","Tier"])

    for i,(sym,r) in enumerate(top_foreign,1):
        ws2.append([
            i,
            sym,
            r.get("foreign_net"),
            r.get("accum_tier")
    ])
    autosize(ws2)
    # =========================================
    # FUNDAMENTAL TOP — HEDGE FUND VERSION
    # =========================================
    

    ws = wb.create_sheet("FUNDAMENTAL_TOP")

    ws.append([
        "Rank","Ticker","FundScore","Quality",
        "ROE","RevenueGrowth","Margin","PE",
        "PBV", "DER", "EPS"
    ])

    header_font = Font(bold=True)
    for c in range(1,12):
        ws.cell(row=1,column=c).font = header_font

    # ranking berdasarkan fund_score
    fund_rank = sorted(
        [(s,r) for s,r in results],
        key=lambda x: x[1].get("fund_score",0),
        reverse=True
    )

    green = PatternFill("solid", fgColor="C6EFCE")
    red   = PatternFill("solid", fgColor="FFC7CE")
    yellow= PatternFill("solid", fgColor="FFF2CC")

    for i,(sym,r) in enumerate(fund_rank[:40],1):

        roe    = r.get("roe") or 0
        growth = r.get("growth") or 0
        margin = r.get("margin") or 0
        pe     = r.get("pe") or 0
        pbv    = r.get("pbv") or 0
        der    = r.get("der") or 0
        eps    = r.get("eps") or 0  

        quality = r.get("fund_quality","MID")

        ws.append([
            i,
            sym,
            r.get("fund_score",0),
            quality,
            roe,
            growth,
            margin,
            pe,
            r.get("pbv"),
            r.get("der"),
            r.get("eps")
        ])

        row = ws.max_row

        # format persen
        ws[f"E{row}"].number_format = '0.0%'
        ws[f"F{row}"].number_format = '0.0%'
        ws[f"G{row}"].number_format = '0.0%'
        ws[f"I{row}"].number_format = '0.00'
        ws[f"J{row}"].number_format = '0.00'
        ws[f"K{row}"].number_format = '0.00'

        # ROE color
        if roe > 0.18:
            ws[f"E{row}"].fill = green
        elif roe < 0:
            ws[f"E{row}"].fill = red

        # growth color
        if growth > 0.15:
            ws[f"F{row}"].fill = green
        elif growth < 0:
            ws[f"F{row}"].fill = red

        # margin color
        if margin > 0.20:
            ws[f"G{row}"].fill = green

        # PE logic
        if pe > 0 and pe < 12:
            ws[f"H{row}"].fill = green
        elif pe > 40:
            ws[f"H{row}"].fill = red

        # MONSTER STOCK highlight
        if roe>0.18 and growth>0.15 and pe<20:
            for col in range(1,9):
                ws.cell(row=row,column=col).fill = yellow
        # PBV
        if pbv > 0 and pbv < 2:
            ws[f"I{row}"].fill = green
        elif pbv > 5:
            ws[f"I{row}"].fill = red
        # DER
        if der < 1:
            ws[f"J{row}"].fill = green
        elif der > 2:
            ws[f"J{row}"].fill = red
        #EPS
        if eps > 0:
            ws[f"K{row}"].fill = green

    autosize(ws)

    wb.save(file)
    print("💼 EXCEL EXPORTED:", file)

def load_foreign_today():
    path = "data/foreign_cache/foreign_today.csv"
    if not os.path.exists(path):
        return 0
    
    df = pd.read_csv(path)
    return df["ForeignNet"].sum()

def accum_tier(net):
    """
    Klasifikasi kekuatan akumulasi foreign.
    """
    if net >= 200_000_000_000:
        return "ULTRA ACCUM"
    elif net >= 50_000_000_000:
        return "STRONG ACCUM"
    elif net >= 5_000_000_000:
        return "ACCUM"
    elif net <= -100_000_000_000:
        return "HEAVY DISTRIB"
    elif net <= -20_000_000_000:
        return "DISTRIB"
    return "NEUTRAL"

def save_watchlist(results):

    os.makedirs("runtime", exist_ok=True)

    watchlist = []

    for sym, r in results[:WATCHLIST_TOPN]:

        watchlist.append({
            "symbol": sym,
            "entry_low": r["entry_low"],
            "entry_high": r["entry_high"],
            "stoploss": r["stoploss"],
            "tp2": r["tp2"],
            "score": r["score"]
        })

    df = pd.DataFrame(watchlist)

    file = "runtime/watchlist.csv"

    df.to_csv(file, index=False)

    print("📡 Watchlist saved:", file)

# ==========================
# MAIN
# ==========================
def run():

    print("\n🏛️ IHSG INSTITUTIONAL SCANNER")
    print("Time:", datetime.now())

    tickers = pd.read_csv("data/universe_institutional.csv", header=None)[0].tolist()
    tickers = tickers[:BATCH_LIMIT]

    market_regime = "UNKNOWN"
    ihsg_trend = "UNKNOWN"

    try:
        ihsg_df = safe_download("^JKSE")
        
        if ihsg_df is None or ihsg_df.empty: 
            print("IHSG Failed, Fallback Empty")
            ihsg_df = pd.DataFrame()
        else:
            market_regime, ihsg_trend = get_market_context(ihsg_df)

    except Exception as e:
        print("IHSG Downloader error", e)
        ihsg_df = pd.DataFrame()

    results = []
    foreign_map = stock_foreign_map()

    print("🌍 Foreign loaded:", len(foreign_map))

    for i, sym in enumerate(tickers):

        if i % 50 == 0:
            print(f"Scanning {i}/{len(tickers)} {sym}")

        # =========================
        # DOWNLOAD
        # =========================
        try:
            guard()
            df = safe_download(sym)
            memory_guard(i)
        except Exception as e:
            print(f"Download error {sym}:", e)
            continue

        if df is None or df.empty or len(df) < 30:
            print(f"Skip {sym} (no data)")
            continue
        avg_vol = df["Volume"].rolling(20).mean().iloc[-1]

        if pd.isna(avg_vol) or avg_vol < 500000:
            print(f"Skip {sym} (low liquidity)")
            continue

            
        # =========================
        # EVALUATE
        # =========================
        try:
            res = evaluate(df, ihsg_df)
        except Exception as e:
            print(f"Eval error {sym}:", e)
            continue

        if not res:
            continue

        res["raw_score"] = float(res["score"])
        res["sector"] = SECTOR_MAP.get(sym, "Unknown")

        # =========================
        # NORMALIZE BEI TICK SIZE
        # =========================

        res["entry_low"]  = floor_tick(res["entry_low"])
        res["entry_high"] = ceil_tick(res["entry_high"])

        res["stoploss"] = floor_tick(res["stoploss"])

        res["tp1"] = ceil_tick(res["tp1"])
        res["tp2"] = ceil_tick(res["tp2"])
        res["tp3"] = ceil_tick(res["tp3"])

        tick = tick_size(res["entry_low"])

        # minimal entry range = 5 tick
        min_range = tick * 4

        recent_high = df["High"].rolling(20).max().iloc[-1]

        if res["entry_high"] - res["entry_low"] < min_range:
            res["entry_high"] = ceil_tick(res["entry_low"] + min_range)
        if res["entry_high"] > recent_high * 0.99:
            res["entry_high"] = floor_tick(recent_high * 0.99)
        if res["tp2"] <= res["entry_high"]:
            res["tp2"] = ceil_tick(res["entry_high"] * 1.05)
        if res["tp1"] <= res["entry_high"]:
            res["tp1"] =  ceil_tick(res["entry_high"] * 1.03)
        if res["tp3"] <= res["tp2"]:
            res["tp3"] = ceil_tick(res["tp2"] * 1.08)
        if res["entry_high"] <= res["entry_low"]:
            res["entry_high"] = ceil_tick(res["entry_low"] + tick * 2)

        # =========================
        # FOREIGN DATA
        # =========================
        base = sym.split(".")[0].upper()
        net = foreign_map.get(base, 0)

        res["foreign_net"] = net
        res["foreign_status"] = classify(net)
        res["accum_tier"] = accum_tier(net)

        # ===== FLOW SCORE =====
        if net >= 200_000_000_000:
            res["raw_score"] += 12
            res["flow_tier"] = "ULTRA"
        elif net >= 50_000_000_000:
            res["raw_score"] += 8
            res["flow_tier"] = "STRONG"
        elif net >= 5_000_000_000:
            res["raw_score"] += 3
            res["flow_tier"] = "ACCUM"
        elif net <= -150_000_000_000:
            res["raw_score"] -= 12
            res["flow_tier"] = "HEAVY SELL"
        elif net <= -30_000_000_000:
            res["raw_score"] -= 6
            res["flow_tier"] = "SELL"
        else:
            res["flow_tier"] = "NEUTRAL"

        # =========================
        # ACCUMULATION DETECTOR
        # =========================
        try:
            high20 = df["High"].rolling(20).max().iloc[-1]
            low10 = df["Low"].rolling(10).min().iloc[-1]
            high10 = df["High"].rolling(10).max().iloc[-1]

            last = df["Close"].iloc[-1]
            vol = df["Volume"].iloc[-1]
            avgvol = df["Volume"].rolling(20).mean().iloc[-1]

            sideways_range = (high10 - low10) / last

            if (
                net > 0 and
                last < high20 * 0.98 and
                vol > avgvol * 1.2 and
                sideways_range < 0.08
            ):
                res["accumulation"] = True
                res["raw_score"] += 5
            else:
                res["accumulation"] = False

        except:
            res["accumulation"] = False

        # =========================
        # EXPECTANCY
        # =========================
        try:
            exp20 = hedge_expectancy(df, horizon=20)
            if exp20:
                res["win20d"], _, _, res["exp20"], res["pf20"] = exp20
        except:
            pass

        # =========================
        # FUNDAMENTAL
        # =========================
        fund = get_fundamental(sym)
        if fund:
            res["fund_score"] = fund.get("fund_score", 0)
            res["fund_quality"] = fund.get("quality", "NEUTRAL")
            res["raw_score"] += fund.get("fund_score", 0) * 0.6

            res["roe"] = fund.get("roe")
            res["growth"] = fund.get("growth")
            res["margin"] = fund.get("margin")
            res["pe"] = fund.get("pe")
            res["der"] = fund.get("der")
            res["pbv"] = fund.get("pbv")
            res["eps"] = fund.get("eps")

        results.append((sym, res))
        
    if not results:
        print("No results")
        return

    top_foreign = get_top_foreign(results)

    raw_scores = [r["raw_score"] for _, r in results]

    for sym, r in results:
        pct = sum(s <= r["raw_score"] for s in raw_scores) / len(raw_scores)
        r["score"] = int(pct * 100)

    results.sort(key=lambda x: x[1]["score"], reverse=True)

    save_watchlist(results)
    
    top_foreign = sorted(
        results,
        key=lambda x: abs(x[1].get("foreign_net", 0)),
        reverse=True
    )[:10]
    # ==========================
    # TOTAL FOREIGN METRICS
    # ==========================
    total_foreign_today = load_foreign_today()
    total_foreign_accum = sum(r.get("foreign_net",0) for _,r in results)

    print("\n🌍 FOREIGN SUMMARY")
    print("TOTAL FOREIGN TODAY :", total_foreign_today)
    print("TOTAL FOREIGN ACCUM :", total_foreign_accum)
    print("\n💰 TOP FOREIGN ACCUMULATION")
    for sym, r in top_foreign:
        net = r["foreign_net"]
        if abs(net) >= 1e9:
            txt = f"{net/1e9:.1f}B"
        elif abs(net) >= 1e6:
            txt = f"{net/1e6:.1f}M"
        else:
            txt = str(int(net))
                
        print(f"{sym}  {txt}  {r['accum_tier']}")

    print("\n📌 TOP WATCHLIST\n")

    for i, (sym, r) in enumerate(results[:WATCHLIST_TOPN], 1):

        print(f"{i}. {sym} Score={r['score']} {star(r['score'])}")
        print(f"   Entry : {r['entry_low']:.0f}-{r['entry_high']:.0f}")
        print(f"   SL    : {r['stoploss']:.0f}")
        print(f"   TP1   : {r['tp1']:.0f}")
        print(f"   TP2   : {r['tp2']:.0f}")
        print(f"   TP3   : {r['tp3']:.0f}")

        net = r["foreign_net"]
        txt = f"{net/1e9:.1f}B" if abs(net)>=1e9 else str(int(net))
        print(f"   🌍 Foreign: {txt} | {r['foreign_status']}\n")
        
        if r.get("accumulation"):
            print("   🟢 Foreign Accumulation Detected")
    
    # ==========================
    # TELEGRAM ALERT
    # ==========================
    try:
        msg = build_telegram_message(results, market_regime, ihsg_trend)
        send(msg)

        top = results[:3]

        for sym, r in top:

            chart = generate_chart(sym, r)
            if chart and os.path.exists(chart):

                caption = f"""
    {sym}
    Score: {r['score']}
    Entry: {int(r['entry_low'])}-{int(r['entry_high'])}
    SL: {int(r['stoploss'])}
    TP: {int(r['tp2'])}

    Win20D: {r.get('win20d',0)*100:.0f}%
    Exp20D: {r.get('exp20',0)*100:.2f}%
    """
                send_photo(chart, caption=caption)
    
        excel_path = "reports/HEDGEFUND_TERMINAL.xlsx"
        
        export_terminal_excel(results, total_foreign_today, top_foreign, market_regime, ihsg_trend)
        
        if os.path.exists(excel_path): 
            send_file(excel_path)
        
        print("Excel Sent to Telegram")

        print("📩 Telegram sent")
    except Exception as e:
        print("Telegram error:", e)

    print("✅ Scan done")

    watchlist = []

    for sym, r in results[:10]:

        watchlist.append({
            "symbol": sym,
            "entry_low": r["entry_low"],
            "entry_high": r["entry_high"],
            "stoploss": r["stoploss"],
            "tp2": r["tp2"],
            "score": r["score"]
        })

    return watchlist


if __name__ == "__main__":
    run()
